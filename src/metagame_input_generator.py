import argparse
import csv
import difflib
import hashlib
import json
import os
import re
import shutil
import sys
import time
import traceback
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass
from dataclasses import field as dataclasses_field
from datetime import UTC, date, datetime, timedelta
from html import unescape
from io import StringIO
from types import SimpleNamespace
from pathlib import Path
from typing import Dict, Iterable, List, Optional
from urllib.parse import urlencode
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl.styles import PatternFill

try:
    from challenge_history_engine import (
        append_to_challenge_history,
        run_challenge_statistics,
        sync_challenge_history_window,
    )
except ImportError:
    try:
        from .challenge_history_engine import (
            append_to_challenge_history,
            run_challenge_statistics,
            sync_challenge_history_window,
        )
    except ImportError:
        append_to_challenge_history = None  # type: ignore[assignment]
        run_challenge_statistics = None  # type: ignore[assignment]
        sync_challenge_history_window = None  # type: ignore[assignment]

try:
    from statistics_engine import run_statistics
except ImportError:
    try:
        from .statistics_engine import run_statistics
    except ImportError:
        run_statistics = None  # type: ignore[assignment]

try:
    from league_engine import run_league_update
except ImportError:
    try:
        from .league_engine import run_league_update
    except ImportError:
        run_league_update = None  # type: ignore[assignment]

try:
    from league_site_export import export_league_site
except ImportError:
    try:
        from .league_site_export import export_league_site
    except ImportError:
        export_league_site = None  # type: ignore[assignment]


API_BASE = "https://api.videreproject.com"
MTGO_BASE = "https://www.mtgo.com"
MTG_GOLDFISH_BASE = "https://www.mtggoldfish.com"
CENSUS_BASE = "https://census.daybreakgames.com"
API_RETRY_ATTEMPTS = 3
API_RETRY_DELAYS_SECONDS = [1, 2, 4]

# Optional fallback chain.
# Preferred source is --backup-url / GUI settings, with MTG_BACKUP_URL left as a
# secondary fallback for CLI/manual use.
# The code will try both:
#   <base>/<endpoint>
#   <base>/api/<endpoint>
# so the configured URL can point either at the host root or the API root.

_CONFIGURED_BACKUP_SERVER_URL: Optional[str] = None


def configure_backup_server_url(value: Optional[str]) -> None:
    global _CONFIGURED_BACKUP_SERVER_URL
    normalized = str(value or "").strip().rstrip("/")
    _CONFIGURED_BACKUP_SERVER_URL = normalized or None


def _get_backup_server_url() -> Optional[str]:
    if _CONFIGURED_BACKUP_SERVER_URL:
        return _CONFIGURED_BACKUP_SERVER_URL

    value = str(os.environ.get("MTG_BACKUP_URL") or "").strip().rstrip("/")
    return value or None


def _candidate_backup_urls(endpoint: str, qs: str) -> List[str]:
    base = _get_backup_server_url()
    if not base:
        return []

    base = base.rstrip("/")
    candidates: List[str] = [f"{base}/{endpoint}?{qs}"]
    if not base.lower().endswith("/api"):
        candidates.append(f"{base}/api/{endpoint}?{qs}")

    seen: set[str] = set()
    deduped: List[str] = []
    for url in candidates:
        if url in seen:
            continue
        seen.add(url)
        deduped.append(url)
    return deduped

# Local JSON cache lives next to outputs/ in the repo root
_CACHE_DIR: Path = Path(__file__).resolve().parent.parent / "outputs" / "cache"


def _cache_key(endpoint: str, params: Dict[str, object]) -> str:
    raw = endpoint + "?" + urlencode(sorted((str(k), str(v)) for k, v in params.items()))
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def _cache_path(endpoint: str, params: Dict[str, object]) -> Path:
    safe_ep = endpoint.replace("/", "_")
    return _CACHE_DIR / f"{safe_ep}_{_cache_key(endpoint, params)}.json"


def _save_cache(path: Path, data: Dict[str, object]) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(data), encoding="utf-8")
    except OSError:
        pass  # cache write failure is non-fatal


def _load_cache(path: Path) -> Optional[Dict[str, object]]:
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        pass
    return None


@dataclass
class ArchetypeRule:
    pattern: str
    archetype: str
    match_type: str = "contains"
    priority: int = 100


@dataclass
class DeckAliasRule:
    pattern: str
    canonical_name: str
    match_type: str = "contains"
    priority: int = 100


@dataclass
class UserDeckMapping:
    raw_name: str
    canonical_name: str
    archetype: str


@dataclass
class GenerationRunResult:
    week_start: date
    week_end: date
    run_dir: Path
    row_count: int
    mapped_from_user: int
    primary_count: int
    fallback_count: int
    imputed_count: int
    rogue_threshold: float
    csv_path: Optional[Path] = None
    xlsx_path: Optional[Path] = None
    rogue_csv_path: Optional[Path] = None
    rogue_xlsx_path: Optional[Path] = None
    rogue_xml_path: Optional[Path] = None
    unknown_output_path: Optional[Path] = None
    alias_suggestions_path: Optional[Path] = None
    challenge_decklist_path: Optional[Path] = None
    challenge_compare_path: Optional[Path] = None
    challenge_event_url: Optional[str] = None
    challenge_history_path: Optional[Path] = None
    premier_history_path: Optional[Path] = None
    challenge_stats_path: Optional[Path] = None
    challenge_events_fetched: int = 0
    challenge_review_items: List[dict] = dataclasses_field(default_factory=list)


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def latest_sunday(reference: date) -> date:
    days_since_sunday = (reference.weekday() + 1) % 7
    return reference - timedelta(days=days_since_sunday)


def parse_percent(value, decimal: bool) -> Optional[float]:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace("%", "").replace(",", ".")
        if text == "":
            return None
        try:
            number = float(text)
        except ValueError:
            return None

    if decimal:
        if number > 1:
            return number / 100.0
        return number
    return number


def parse_count(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return max(0, int(value))

    text = str(value).strip().replace(",", "")
    if text == "":
        return 0
    try:
        return max(0, int(float(text)))
    except ValueError:
        return 0


def _api_fetch_once(url: str) -> Dict[str, object]:
    """Single HTTP GET, no retries. Used for backup candidates."""
    req = Request(
        url,
        headers={
            "Accept": "application/json",
            "User-Agent": "MTG-Metagame-Analyzer/1.0 (+https://github.com)",
        },
    )
    try:
        with urlopen(req, timeout=15) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as err:
        raise RuntimeError(f"HTTP {err.code}: {url}") from err
    except URLError as err:
        raise RuntimeError(f"Connection error: {err}") from err


def _api_fetch(url: str) -> Dict[str, object]:
    """Single HTTP GET with retries. Always raises RuntimeError on final failure."""
    req = Request(
        url,
        headers={
            "Accept": "application/json",
            "User-Agent": "MTG-Metagame-Analyzer/1.0 (+https://github.com)",
        },
    )
    for attempt in range(API_RETRY_ATTEMPTS):
        try:
            with urlopen(req, timeout=15) as response:
                return json.loads(response.read().decode("utf-8"))
        except HTTPError as err:
            if err.code >= 500 and attempt < API_RETRY_ATTEMPTS - 1:
                delay = API_RETRY_DELAYS_SECONDS[min(attempt, len(API_RETRY_DELAYS_SECONDS) - 1)]
                print(
                    f"[WARN] HTTP {err.code} on '{url}'. "
                    f"Retrying in {delay}s ({attempt + 1}/{API_RETRY_ATTEMPTS})..."
                )
                time.sleep(delay)
                continue
            # Convert all HTTP errors (including 4xx) to RuntimeError so callers
            # can rely on a single exception type for fallback decisions.
            raise RuntimeError(f"HTTP {err.code}: {url}") from err
        except URLError as err:
            if attempt < API_RETRY_ATTEMPTS - 1:
                delay = API_RETRY_DELAYS_SECONDS[min(attempt, len(API_RETRY_DELAYS_SECONDS) - 1)]
                print(
                    f"[WARN] Connection error on '{url}'. "
                    f"Retrying in {delay}s ({attempt + 1}/{API_RETRY_ATTEMPTS})..."
                )
                time.sleep(delay)
                continue
            raise RuntimeError(f"Connection error: {err}") from err
    raise RuntimeError(f"All retries exhausted for: {url}")


def api_get(endpoint: str, params: Dict[str, object]) -> Dict[str, object]:
    """Fetch from API with fallback chain: primary API → backup VPS → local cache."""
    qs = urlencode(params)
    cache_path = _cache_path(endpoint, params)

    # 1. Try primary API
    try:
        data = _api_fetch(f"{API_BASE}/{endpoint}?{qs}")
        _save_cache(cache_path, data)  # update local cache on every success
        return data
    except (RuntimeError, HTTPError, URLError) as primary_err:
        print(f"[WARN] Primary API failed for '{endpoint}': {primary_err}")

    # 2. Try backup VPS (if configured)
    backup_candidates = _candidate_backup_urls(endpoint, qs)
    if backup_candidates:
        for backup_url in backup_candidates:
            try:
                data = _api_fetch_once(backup_url)
                print(f"[INFO] Using backup server data for '{endpoint}' via {backup_url}.")
                _save_cache(cache_path, data)  # keep local cache current
                return data
            except (RuntimeError, HTTPError, URLError) as backup_err:
                print(f"[WARN] Backup server failed for '{endpoint}' via {backup_url}: {backup_err}")

    # 3. Fall back to local cache
    cached = _load_cache(cache_path)
    if cached is not None:
        print(f"[INFO] Using local cache for '{endpoint}' ({cache_path.name}).")
        return cached

    raise RuntimeError(
        "API temporary error. Please try again later. "
        f"Endpoint='{endpoint}'. No backup server or local cache available."
    )


def _fetch_text(url: str, timeout: int = 45) -> str:
    req = Request(
        url,
        headers={
            "Accept": "text/html,application/xhtml+xml",
            "User-Agent": "MTG-Metagame-Input-Generator/1.0 (+https://github.com)",
        },
    )
    with urlopen(req, timeout=timeout) as response:
        return response.read().decode("utf-8", errors="replace")


def _strip_html(text: str) -> str:
    plain = re.sub(r"<[^>]+>", " ", str(text))
    plain = unescape(plain)
    plain = re.sub(r"\s+", " ", plain).strip()
    return plain


def _format_slug_for_url(format_name: str) -> str:
    return normalize_name(format_name).replace(" ", "-")


def _normalize_format_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_name(value))


def _calendar_rows_for_window(start_date: date, end_date: date, page_limit: int = 1000) -> List[dict]:
    start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=UTC)
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time()).replace(tzinfo=UTC) - timedelta(
        milliseconds=1
    )
    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)

    rows: List[dict] = []
    offset = 0
    while True:
        params = {
            "c:limit": page_limit,
            "c:start": offset,
            "starttime": f"]{start_ms}",
            "endtime": f"[{end_ms}",
        }
        url = f"{CENSUS_BASE}/s:dgc/get/mtgo/tournament_calendar?{urlencode(params)}"
        payload = _api_fetch_once(url)
        batch = payload.get("tournament_calendar_list")
        if not isinstance(batch, list) or not batch:
            break
        rows.extend(batch)
        if len(batch) < page_limit:
            break
        offset += len(batch)

    return rows


def _find_challenges_from_decklists_index(
    format_name: str, start_date: date, end_date: date
) -> List[dict]:
    html = _fetch_text(f"{MTGO_BASE}/decklists")
    format_slug = _format_slug_for_url(format_name)

    pattern = re.compile(
        rf'href="/decklist/({re.escape(format_slug)}-challenge-(\d+)-(\d{{4}}-\d{{2}}-\d{{2}})(\d+))"',
        flags=re.IGNORECASE,
    )

    seen: dict = {}
    for match in pattern.finditer(html):
        slug = match.group(1)
        challenge_size = int(match.group(2))
        event_date_str = match.group(3)
        numeric_id = int(match.group(4))

        try:
            ev_date = date.fromisoformat(event_date_str)
        except ValueError:
            continue

        if not (start_date <= ev_date <= end_date):
            continue

        key = (event_date_str, challenge_size)
        if key not in seen or numeric_id > seen[key]["_numeric_id"]:
            seen[key] = {
                "slug": slug,
                "challenge_size": challenge_size,
                "event_date": event_date_str,
                "mtgo_url": f"{MTGO_BASE}/decklist/{slug}",
                "_numeric_id": numeric_id,
            }

    results: List[dict] = []
    for info in seen.values():
        clean = {k: v for k, v in info.items() if k != "_numeric_id"}
        results.append(clean)
    results.sort(key=lambda x: (x["event_date"], x["challenge_size"]), reverse=True)
    return results


def _find_challenges_in_window(
    format_name: str, start_date: date, end_date: date
) -> List[dict]:
    """Return all MTGO Challenges (any size) for *format_name* whose event_date
    falls within the closed interval [start_date, end_date].

    A dict per challenge is returned, sorted by (event_date DESC, challenge_size DESC).
    When the same (event_date, challenge_size) pair appears with different numeric IDs
    (reruns / corrections on MTGO) the highest numeric_id wins.
    """
    format_slug = _format_slug_for_url(format_name)
    normalized_target_format = _normalize_format_token(format_name)
    challenge_desc_re = re.compile(r"^(?P<fmt>.+?)\s+challenge(?:\s+(?P<size>\d+))?$", flags=re.IGNORECASE)

    try:
        calendar_rows = _calendar_rows_for_window(start_date, end_date)
    except Exception as err:
        print(f"[WARN] MTGO calendar fetch failed for challenge discovery: {err}. Falling back to /decklists index.")
        return _find_challenges_from_decklists_index(format_name, start_date, end_date)

    seen: Dict[tuple[str, int, str], dict] = {}
    for row in calendar_rows:
        description = str(row.get("description") or "").strip()
        match = challenge_desc_re.match(description)
        if not match:
            continue

        row_format = _normalize_format_token(match.group("fmt") or "")
        if row_format != normalized_target_format:
            continue

        size_raw = match.group("size") or row.get("minimumplayers")
        try:
            challenge_size = int(str(size_raw).strip())
        except (TypeError, ValueError):
            continue
        if challenge_size <= 0:
            continue

        starttime = str(row.get("starttime") or "").strip()
        if len(starttime) < 10:
            continue
        event_date_str = starttime[:10]
        try:
            ev_date = date.fromisoformat(event_date_str)
        except ValueError:
            continue
        if not (start_date <= ev_date <= end_date):
            continue

        tournament_id_raw = str(row.get("tournamentid") or "").strip()
        if not tournament_id_raw.isdigit():
            continue
        numeric_id = int(tournament_id_raw)
        slug = f"{format_slug}-challenge-{challenge_size}-{event_date_str}{numeric_id}"

        key = (event_date_str, challenge_size, tournament_id_raw)
        if key not in seen:
            seen[key] = {
                "slug": slug,
                "challenge_size": challenge_size,
                "event_date": event_date_str,
                "mtgo_url": f"{MTGO_BASE}/decklist/{slug}",
            }

    results: List[dict] = []
    for info in seen.values():
        results.append(info)
    results.sort(key=lambda x: (x["event_date"], x["challenge_size"]), reverse=True)

    if results:
        return results

    print("[WARN] MTGO calendar returned no matching challenge events; falling back to /decklists index.")
    return _find_challenges_from_decklists_index(format_name, start_date, end_date)


def _find_latest_mtgo_challenge(format_name: str) -> Optional[dict]:
    """Legacy wrapper – returns the single most recent challenge in the past 30 days."""
    results = _find_challenges_in_window(
        format_name,
        start_date=date.today() - timedelta(days=30),
        end_date=date.today(),
    )
    return results[0] if results else None


def _build_mtggoldfish_challenge_url(format_name: str, challenge_size: int, event_date: str) -> str:
    format_slug = _format_slug_for_url(format_name)
    return f"{MTG_GOLDFISH_BASE}/tournament/{format_slug}-challenge-{challenge_size}-{event_date}#online"


def _parse_mtggoldfish_challenge_rows(html: str) -> List[dict]:
    rows: List[dict] = []
    row_pattern = re.compile(r"<tr[^>]*>(.*?)</tr>", flags=re.IGNORECASE | re.DOTALL)
    cell_pattern = re.compile(r"<td[^>]*>(.*?)</td>", flags=re.IGNORECASE | re.DOTALL)

    for row_html in row_pattern.findall(html):
        cells_html = cell_pattern.findall(row_html)
        if len(cells_html) < 3:
            continue

        raw_place = _strip_html(cells_html[0])
        place_match = re.fullmatch(r"(\d+)(?:st|nd|rd|th)", raw_place.lower())
        if not place_match:
            continue

        raw_deck = _strip_html(cells_html[1])
        raw_deck = re.sub(r"\s*mana cost\s*:.*$", "", raw_deck, flags=re.IGNORECASE)
        raw_deck = raw_deck.strip()
        pilot = _strip_html(cells_html[2])

        if not raw_deck or not pilot:
            continue

        rows.append(
            {
                "Place": int(place_match.group(1)),
                "Deck": raw_deck,
                "Pilot": pilot,
            }
        )

    rows.sort(key=lambda item: item["Place"])
    return rows


def fetch_latest_challenge_decks_from_mtggoldfish(format_name: str) -> tuple[dict, pd.DataFrame]:
    """Legacy wrapper: fetch the single most-recent Challenge. Uses _parse_single_challenge internally."""
    latest = _find_latest_mtgo_challenge(format_name)
    if not latest:
        raise RuntimeError(f"No recent {format_name} Challenge found on MTGO decklists index.")
    return _parse_single_challenge(latest, format_name)


def _parse_single_challenge(
    challenge_info: dict, format_name: str
) -> tuple[dict, pd.DataFrame]:
    """Fetch and parse one challenge from MTGGoldfish given its MTGO info dict.

    Returns (event_info, df) where df has columns: Place, Deck, Pilot.
    Raises RuntimeError if parsing fails.
    """
    goldfish_url = _build_mtggoldfish_challenge_url(
        format_name=format_name,
        challenge_size=int(challenge_info["challenge_size"]),
        event_date=str(challenge_info["event_date"]),
    )
    html = _fetch_text(goldfish_url)
    parsed_rows = _parse_mtggoldfish_challenge_rows(html)

    if not parsed_rows:
        # Fallback: pandas read_html
        try:
            candidate_tables = pd.read_html(StringIO(html))
        except ValueError:
            candidate_tables = []
        for table in candidate_tables:
            lowered = [str(col).strip().lower() for col in table.columns]
            if not {"place", "deck", "pilot"}.issubset(set(lowered)):
                continue
            renamed = table.rename(columns={
                table.columns[lowered.index("place")]: "Place",
                table.columns[lowered.index("deck")]: "Deck",
                table.columns[lowered.index("pilot")]: "Pilot",
            })
            rows: List[dict] = []
            for _, row in renamed.iterrows():
                place_raw = str(row.get("Place") or "").strip().lower()
                place_match = re.fullmatch(r"(\d+)(?:st|nd|rd|th)", place_raw)
                if not place_match:
                    continue
                deck_name = re.sub(
                    r"\s*mana cost\s*:.*$", "",
                    str(row.get("Deck") or "").strip(), flags=re.IGNORECASE,
                ).strip()
                pilot_name = str(row.get("Pilot") or "").strip()
                if not deck_name or not pilot_name:
                    continue
                rows.append({"Place": int(place_match.group(1)), "Deck": deck_name, "Pilot": pilot_name})
            if rows:
                parsed_rows = sorted(rows, key=lambda item: item["Place"])
                break

    if not parsed_rows:
        raise RuntimeError(f"Could not parse deck rows from MTGGoldfish: {goldfish_url}")

    df = pd.DataFrame(parsed_rows)
    df = df.drop_duplicates(subset=["Place", "Deck", "Pilot"]).sort_values("Place").reset_index(drop=True)
    event_info = {
        "format": format_name,
        "event_date": str(challenge_info["event_date"]),
        "challenge_size": int(challenge_info["challenge_size"]),
        "mtgo_url": str(challenge_info["mtgo_url"]),
        "mtggoldfish_url": goldfish_url,
    }
    return event_info, df


def roster_content_signature(df: pd.DataFrame) -> str:
    """Order-independent identity of a Challenge roster: hash of sorted (pilot, deck) pairs.

    Two events with the same signature had the identical set of pilots/decks, which is how we
    detect MTGO republishing the same tournament under a second tournament_id (same MTGGoldfish
    page gets scraped twice) rather than trusting the discovery-layer (date, size) key alone.
    """
    pairs = sorted(
        (normalize_name(p), normalize_name(d))
        for p, d in zip(df.get("Pilot", []), df.get("Deck", []))
    )
    canonical = "|".join(f"{p}::{d}" for p, d in pairs)
    return hashlib.sha1(canonical.encode("utf-8")).hexdigest()[:12]


def fetch_challenges_in_window_from_mtggoldfish(
    format_name: str,
    start_date: date,
    end_date: date,
    aliases: Optional[List[DeckAliasRule]] = None,
    rules: Optional[List[ArchetypeRule]] = None,
    challenge_mappings: Optional[List[UserDeckMapping]] = None,
) -> List[tuple[dict, pd.DataFrame]]:
    """Fetch ALL challenges (C32 and C64) within [start_date, end_date] from MTGGoldfish.

    Returns a list of (event_info, df) tuples, one per event.
    Each df has columns: Place, Deck, Archetype, Pilot.
    If *aliases* and *rules* are provided, Deck names are canonicalised and
    Archetype is assigned; otherwise they pass through raw.
    """
    challenges = _find_challenges_in_window(format_name, start_date, end_date)
    if not challenges:
        raise RuntimeError(
            f"No {format_name} Challenges found on MTGO between "
            f"{start_date} and {end_date}."
        )

    results: List[tuple[dict, pd.DataFrame]] = []
    mapping_lookup_exact: Dict[str, UserDeckMapping] = {}
    for mapping in challenge_mappings or []:
        key = normalize_name(mapping.raw_name)
        if key and key not in mapping_lookup_exact:
            mapping_lookup_exact[key] = mapping

    parsed: List[tuple[dict, pd.DataFrame]] = []
    for info in challenges:
        try:
            event_info, df = _parse_single_challenge(info, format_name)
        except Exception:
            continue  # skip events that fail to parse; caller gets the rest
        parsed.append((event_info, df))

    # Dedup by roster content, grouped by (event_date, challenge_size). MTGO sometimes lists the
    # same tournament twice under different tournament_ids; since the MTGGoldfish scrape URL only
    # depends on (format, size, date), both entries fetch the identical page. A content-signature
    # mismatch within a group means a genuine same-day/same-tier collision MTGGoldfish itself can't
    # disambiguate (known source limitation) -- still counted as one event, but flagged.
    groups: Dict[tuple[str, int], List[tuple[dict, pd.DataFrame]]] = defaultdict(list)
    for event_info, df in parsed:
        key = (str(event_info["event_date"]), int(event_info["challenge_size"]))
        groups[key].append((event_info, df))

    deduped: List[tuple[dict, pd.DataFrame]] = []
    for (event_date_str, challenge_size), entries in groups.items():
        if len(entries) == 1:
            deduped.append(entries[0])
            continue

        entries_sorted = sorted(entries, key=lambda item: str(item[0].get("slug") or ""))
        signatures = [roster_content_signature(df) for _, df in entries_sorted]
        kept_event_info, kept_df = entries_sorted[0]

        if len(set(signatures)) == 1:
            print(
                f"[challenge-dedup] Merged {len(entries_sorted)} duplicate publication(s) of "
                f"C{challenge_size} {event_date_str} (identical roster, MTGO republish)"
            )
        else:
            print(
                f"[WARN] [challenge-dedup] C{challenge_size} {event_date_str} has "
                f"{len(entries_sorted)} same-day/same-tier entries with differing rosters; "
                "MTGGoldfish cannot disambiguate them (known source limitation) -- counted as 1 event."
            )
        deduped.append((kept_event_info, kept_df))

    for event_info, df in deduped:
        # First pass: explicit challenge-only overrides (separate history/mapping space).
        if mapping_lookup_exact:
            def _apply_challenge_mapping(raw_name: str) -> tuple[str, Optional[str]]:
                key = normalize_name(str(raw_name))
                mapping = mapping_lookup_exact.get(key)
                if mapping is None:
                    return str(raw_name), None
                return mapping.canonical_name, mapping.archetype

            mapped_values = df["Deck"].apply(_apply_challenge_mapping)
            df["Deck"] = mapped_values.apply(lambda x: x[0])
            df["Archetype"] = mapped_values.apply(lambda x: x[1] if x[1] else "")
        else:
            df["Archetype"] = ""

        if aliases is not None:
            df["Deck"] = df["Deck"].apply(
                lambda name: canonicalize_deck_name(str(name), aliases)
            )
        if rules is not None:
            df["Archetype"] = df.apply(
                lambda row: str(row.get("Archetype") or "").strip()
                or classify_archetype(str(row.get("Deck") or ""), rules),
                axis=1,
            )
        else:
            df["Archetype"] = df.apply(
                lambda row: str(row.get("Archetype") or "").strip()
                or heuristic_archetype(str(row.get("Deck") or "")),
                axis=1,
            )

        results.append((event_info, df))

    return results


def build_challenge_vs_metagame_report(
    challenge_df: pd.DataFrame,
    metagame_df: pd.DataFrame,
    aliases: List[DeckAliasRule],
    rules: List[ArchetypeRule],
) -> pd.DataFrame:
    if challenge_df.empty:
        return pd.DataFrame(
            columns=[
                "Deck",
                "Archetype",
                "Challenge Count",
                "Challenge Share %",
                "Meta Share %",
                "Delta pp",
            ]
        )

    challenge_norm = challenge_df.copy()
    challenge_norm["Deck"] = challenge_norm["Deck"].apply(lambda name: canonicalize_deck_name(str(name), aliases))
    challenge_summary = (
        challenge_norm.groupby("Deck", as_index=False)
        .size()
        .rename(columns={"size": "Challenge Count"})
    )
    total_decks = int(challenge_summary["Challenge Count"].sum())
    challenge_summary["Challenge Share %"] = (
        (challenge_summary["Challenge Count"] / max(total_decks, 1)) * 100.0
    ).round(2)

    meta_summary = aggregate_by_deck(metagame_df)
    if "Deck" not in meta_summary.columns:
        meta_summary = pd.DataFrame(columns=["Deck", "Meta"])
    meta_summary = meta_summary[["Deck", "Meta"]].copy()
    meta_summary["Meta Share %"] = pd.to_numeric(meta_summary["Meta"], errors="coerce").fillna(0).round(2)
    meta_summary = meta_summary.drop(columns=["Meta"])

    merged = challenge_summary.merge(meta_summary, on="Deck", how="outer")
    merged["Challenge Count"] = pd.to_numeric(merged["Challenge Count"], errors="coerce").fillna(0).astype(int)
    merged["Challenge Share %"] = pd.to_numeric(merged["Challenge Share %"], errors="coerce").fillna(0).round(2)
    merged["Meta Share %"] = pd.to_numeric(merged["Meta Share %"], errors="coerce").fillna(0).round(2)
    merged["Delta pp"] = (merged["Challenge Share %"] - merged["Meta Share %"]).round(2)
    merged["Archetype"] = merged["Deck"].apply(lambda deck: classify_archetype(str(deck), rules))

    merged = merged[
        [
            "Deck",
            "Archetype",
            "Challenge Count",
            "Challenge Share %",
            "Meta Share %",
            "Delta pp",
        ]
    ].sort_values(["Challenge Count", "Challenge Share %", "Meta Share %"], ascending=[False, False, False])
    return merged.reset_index(drop=True)


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", str(name).strip()).lower()


def _expand_camel_case(name: str) -> str:
    """Expand CamelCase names that have no spaces: IzzetControl → Izzet Control."""
    if " " not in name:
        expanded = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", name)
        if expanded != name:
            return expanded
    return name


def load_rules(path: Path) -> List[ArchetypeRule]:
    if not path.exists():
        return []

    rules: List[ArchetypeRule] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            pattern = (row.get("pattern") or "").strip()
            archetype = (row.get("archetype") or "").strip()
            if not pattern or not archetype:
                continue
            rules.append(
                ArchetypeRule(
                    pattern=pattern,
                    archetype=archetype,
                    match_type=(row.get("match_type") or "contains").strip().lower(),
                    priority=int((row.get("priority") or "100").strip()),
                )
            )
    return sorted(rules, key=lambda rule: rule.priority)


def load_aliases(path: Path) -> List[DeckAliasRule]:
    if not path.exists():
        return []

    aliases: List[DeckAliasRule] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            pattern = (row.get("pattern") or "").strip()
            canonical_name = (row.get("canonical_name") or "").strip()
            if not pattern or not canonical_name:
                continue
            aliases.append(
                DeckAliasRule(
                    pattern=pattern,
                    canonical_name=canonical_name,
                    match_type=(row.get("match_type") or "contains").strip().lower(),
                    priority=int((row.get("priority") or "100").strip()),
                )
            )
    return sorted(aliases, key=lambda alias: alias.priority)


def load_user_deck_mappings(path: Path) -> List[UserDeckMapping]:
    if not path.exists():
        return []

    mappings: List[UserDeckMapping] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            raw_name = (row.get("raw_name") or "").strip()
            canonical_name = (row.get("canonical_name") or "").strip()
            archetype = (row.get("archetype") or "").strip()
            if not raw_name or not canonical_name:
                continue
            mappings.append(
                UserDeckMapping(
                    raw_name=raw_name,
                    canonical_name=canonical_name,
                    archetype=archetype,
                )
            )
    return mappings


def mapping_lookup(mappings: List[UserDeckMapping]) -> Dict[str, UserDeckMapping]:
    return {normalize_name(m.raw_name): m for m in mappings}


def canonicalize_deck_name(deck_name: str, aliases: Iterable[DeckAliasRule]) -> str:
    base_name = re.sub(r"\s+", " ", str(deck_name).strip())
    base_name = re.sub(r"(?i)generic", " ", base_name)
    base_name = re.sub(r"\s+", " ", base_name).strip()
    base_name = _expand_camel_case(base_name)
    deck_norm = normalize_name(base_name)

    for alias in aliases:
        pattern = alias.pattern
        expanded_pattern = _expand_camel_case(pattern)
        mtype = alias.match_type

        if mtype == "exact" and deck_norm == normalize_name(expanded_pattern):
            return alias.canonical_name
        if mtype == "contains" and normalize_name(expanded_pattern) in deck_norm:
            return alias.canonical_name
        if mtype == "regex" and re.search(pattern, base_name, flags=re.IGNORECASE):
            return alias.canonical_name

    if "prowess" in deck_norm:
        return "Prowess"

    return base_name


def heuristic_archetype(deck_name: str) -> str:
    normalized = normalize_name(deck_name)
    patterns = [
        (r"\b(burn|zoo|prowess|aggro|humans|goblins|elves|merfolk|infect|affinity|hammer)\b", "Aggro"),
        (r"\b(midrange|rock|jund|abzan|grixis midrange|golgari)\b", "Midrange"),
        (r"\b(control|azorius|jeskai control|dimir control|uw control|esper control)\b", "Control"),
        (r"\b(combo|storm|belcher|twin|neoform|oops|creativity|ad nauseam|yawgmoth|amulet)\b", "Combo"),
        (r"\b(ramp|tron|titan|eldrazi ramp|scapeshift)\b", "Ramp"),
        (r"\b(tempo|delver|murktide)\b", "Midrange"),
        (r"\b(prison|lantern|lock)\b", "Prison"),
        (r"\b(reanimator|goryo|living end|dredge)\b", "Graveyard"),
    ]
    for pattern, archetype in patterns:
        if re.search(pattern, normalized, flags=re.IGNORECASE):
            return archetype
    return "Unknown"


def classify_archetype(deck_name: str, rules: Iterable[ArchetypeRule]) -> str:
    deck_norm = normalize_name(deck_name)
    for rule in rules:
        pattern = rule.pattern
        mtype = rule.match_type

        if mtype == "exact" and deck_norm == normalize_name(pattern):
            return rule.archetype
        if mtype == "contains" and normalize_name(pattern) in deck_norm:
            return rule.archetype
        if mtype == "regex" and re.search(pattern, deck_name, flags=re.IGNORECASE):
            return rule.archetype

    return heuristic_archetype(deck_name)


def normalize_archetype_label(archetype: str) -> str:
    text = str(archetype or "").strip()
    if normalize_name(text) == "reanimator":
        return "Graveyard"
    return text if text else "Unknown"


def is_code_like_deck_name(deck_name: str) -> bool:
    text = str(deck_name or "").strip()
    if text == "":
        return False
    if re.fullmatch(r"[A-Z]{1,5}", text):
        return True
    norm = normalize_name(text)
    if re.fullmatch(r"[wubrg]{1,5}", norm):
        return True
    return False


def apply_archetype_overrides(raw_deck_name: str, deck_name: str, meta_share: Optional[float], archetype: str) -> str:
    raw_norm = normalize_name(raw_deck_name)
    deck_norm = normalize_name(deck_name)

    if raw_norm == "sultai midrange" or deck_norm == "sultai midrange":
        return "Midrange"

    return archetype


def extract_my_deck_matchups(
    format_name: str,
    my_deck: str,
    start_date: date,
    end_date: date,
    limit: int,
    aliases: List[DeckAliasRule],
) -> Dict[str, tuple[float, int]]:
    payload = api_get(
        "matchups",
        {
            "format": format_name,
            "min_date": start_date.isoformat(),
            "max_date": end_date.isoformat(),
            "limit": limit,
        },
    )

    data = payload.get("data") or []
    if not isinstance(data, list):
        return {}

    my_canonical = canonicalize_deck_name(my_deck, aliases)
    my_norm = normalize_name(my_canonical)
    exact = None
    contains = None

    for row in data:
        archetype = row.get("archetype")
        if not archetype:
            continue
        row_canonical = canonicalize_deck_name(str(archetype), aliases)
        row_norm = normalize_name(row_canonical)
        if row_norm == my_norm:
            exact = row
            break
        if my_norm in row_norm or row_norm in my_norm:
            contains = row

    selected = exact or contains
    if not selected:
        return {}

    grouped_wr: Dict[str, List[tuple[float, int]]] = defaultdict(list)
    for matchup in selected.get("matchups", []):
        opponent = matchup.get("archetype")
        if not opponent:
            continue
        opponent_canonical = canonicalize_deck_name(str(opponent), aliases)
        wr = parse_percent(matchup.get("game_winrate"), decimal=True)
        if wr is None:
            continue
        games = parse_count(matchup.get("game_count"))
        grouped_wr[normalize_name(opponent_canonical)].append((wr, games))

    # If multiple source matchups collapse into one canonical deck family
    # (e.g. Izzet Prowess + Izzet Steel-Cutter -> Prowess), use their mean.
    out: Dict[str, tuple[float, int]] = {}
    for opponent_norm, wr_values in grouped_wr.items():
        total_games = sum(g for _, g in wr_values)
        if total_games > 0:
            weighted_wr = sum(wr * g for wr, g in wr_values) / total_games
            out[opponent_norm] = (float(weighted_wr), int(total_games))
        else:
            out[opponent_norm] = (float(sum(wr for wr, _ in wr_values) / len(wr_values)), 0)
    return out


def fetch_available_decks(
    format_name: str,
    start_date: date,
    end_date: date,
    limit: int,
) -> List[str]:
    payload = api_get(
        "metagame",
        {
            "format": format_name,
            "min_date": start_date.isoformat(),
            "max_date": end_date.isoformat(),
            "limit": limit,
        },
    )

    rows = payload.get("data") or []
    if not isinstance(rows, list):
        return []

    decks: List[str] = []
    seen = set()
    for row in rows:
        name = str(row.get("archetype") or "").strip()
        if not name:
            continue
        key = normalize_name(name)
        if key in seen:
            continue
        seen.add(key)
        decks.append(name)
    return decks


def build_dataset(
    format_name: str,
    week_start: date,
    week_end: date,
    my_deck: str,
    my_window_days: int,
    my_fallback_window_days: int,
    rogue_threshold: float,
    rules: List[ArchetypeRule],
    aliases: List[DeckAliasRule],
    user_mapping: Dict[str, UserDeckMapping],
    metagame_limit: int,
    matchup_limit: int,
) -> tuple[pd.DataFrame, int, Dict[str, int]]:
    metagame_payload = api_get(
        "metagame",
        {
            "format": format_name,
            "min_date": week_start.isoformat(),
            "max_date": week_end.isoformat(),
            "limit": metagame_limit,
        },
    )
    metagame_data = metagame_payload.get("data") or []
    if not isinstance(metagame_data, list) or not metagame_data:
        raise RuntimeError("Brak danych metagame dla zadanego tygodnia.")

    my_start = week_end - timedelta(days=max(1, my_window_days))
    my_wr_primary = extract_my_deck_matchups(
        format_name=format_name,
        my_deck=my_deck,
        start_date=my_start,
        end_date=week_end,
        limit=matchup_limit,
        aliases=aliases,
    )

    my_wr_fallback: Dict[str, tuple[float, int]] = {}
    if my_fallback_window_days > my_window_days:
        fallback_start = week_end - timedelta(days=max(1, my_fallback_window_days))
        my_wr_fallback = extract_my_deck_matchups(
            format_name=format_name,
            my_deck=my_deck,
            start_date=fallback_start,
            end_date=week_end,
            limit=matchup_limit,
            aliases=aliases,
        )

    rows = []
    mapped_from_user = 0
    for entry in metagame_data:
        raw_deck_name = str(entry.get("archetype") or "").strip()
        user_map = user_mapping.get(normalize_name(raw_deck_name))
        if user_map:
            mapped_from_user += 1

        canonical_source = user_map.canonical_name if user_map else raw_deck_name
        deck_name = canonicalize_deck_name(canonical_source, aliases)
        if not deck_name:
            continue

        meta_share = parse_percent(entry.get("percentage"), decimal=False)
        winrate = parse_percent(entry.get("game_winrate"), decimal=True)
        archetype_raw = user_map.archetype if (user_map and user_map.archetype) else classify_archetype(deck_name, rules)
        archetype = normalize_archetype_label(archetype_raw)
        archetype = apply_archetype_overrides(raw_deck_name, deck_name, meta_share, archetype)

        deck_lookup_name = deck_name

        deck_norm = normalize_name(deck_lookup_name)
        if deck_norm in my_wr_primary:
            my_wr, my_wr_games = my_wr_primary[deck_norm]
            my_wr_source = "primary"
        elif deck_norm in my_wr_fallback:
            my_wr, my_wr_games = my_wr_fallback[deck_norm]
            my_wr_source = "fallback"
        else:
            my_wr = pd.NA
            my_wr_games = 0
            my_wr_source = "none"

        rows.append(
            {
                "Raw Deck": raw_deck_name,
                "Deck": deck_name,
                "Meta": meta_share,
                "Winrate": winrate,
                "Winrate Game Count": parse_count(entry.get("game_count")),
                "Archetype": archetype,
                "My Deck Winrate": my_wr,
                "My Deck Winrate Game Count": my_wr_games,
                "My Deck Winrate Source": my_wr_source,
            }
        )

    df_raw = pd.DataFrame(rows)
    df_raw = df_raw.dropna(subset=["Deck", "Meta", "Winrate"]).copy()
    if df_raw.empty:
        raise RuntimeError("Brak poprawnych rekordow metagame po normalizacji.")

    dup_counts = df_raw.groupby("Deck").size()
    dup_decks = dup_counts[dup_counts > 1]
    if not dup_decks.empty:
        sample = ", ".join(dup_decks.index[:8])
        print(
            "[WARN] API returned duplicate metagame rows for canonical decks: "
            f"{sample}. Aggregating duplicates by deck."
        )

    grouped_rows: List[Dict[str, object]] = []
    for deck_name, grp in df_raw.groupby("Deck", sort=False):
        grp = grp.copy()
        grp["Winrate Game Count"] = pd.to_numeric(grp["Winrate Game Count"], errors="coerce").fillna(0).astype(int)
        grp["Winrate"] = pd.to_numeric(grp["Winrate"], errors="coerce")
        grp["My Deck Winrate"] = pd.to_numeric(grp["My Deck Winrate"], errors="coerce")

        games_total = int(grp["Winrate Game Count"].sum())
        if games_total > 0:
            weighted_winrate = float((grp["Winrate"] * grp["Winrate Game Count"]).sum() / games_total)
        else:
            weighted_winrate = float(grp["Winrate"].mean())

        meta_share = float(pd.to_numeric(grp["Meta"], errors="coerce").fillna(0).sum())

        sources = grp["My Deck Winrate Source"].astype(str)
        if (sources == "primary").any():
            my_wr_source = "primary"
        elif (sources == "fallback").any():
            my_wr_source = "fallback"
        else:
            my_wr_source = "none"

        if my_wr_source == "primary":
            wr_slice = grp.loc[sources == "primary", "My Deck Winrate"]
            wr_games_slice = grp.loc[sources == "primary", "My Deck Winrate Game Count"]
        elif my_wr_source == "fallback":
            wr_slice = grp.loc[sources == "fallback", "My Deck Winrate"]
            wr_games_slice = grp.loc[sources == "fallback", "My Deck Winrate Game Count"]
        else:
            wr_slice = pd.Series([], dtype=float)
            wr_games_slice = pd.Series([], dtype=float)

        my_wr_value = float(wr_slice.dropna().mean()) if not wr_slice.dropna().empty else pd.NA
        my_wr_games = int(pd.to_numeric(wr_games_slice, errors="coerce").fillna(0).max()) if not wr_games_slice.empty else 0

        archetypes = [normalize_archetype_label(str(a)) for a in grp["Archetype"].tolist()]
        archetype = pd.Series(archetypes).mode().iloc[0] if archetypes else "Unknown"

        source_names = sorted({str(v).strip() for v in grp["Raw Deck"].tolist() if str(v).strip()})
        grouped_rows.append(
            {
                "Deck": deck_name,
                "Source Deck Names": " | ".join(source_names),
                "Meta": meta_share,
                "Winrate": weighted_winrate,
                "Winrate Game Count": games_total,
                "Archetype": archetype,
                "My Deck Winrate": my_wr_value,
                "My Deck Winrate Game Count": my_wr_games,
                "My Deck Winrate Source": my_wr_source,
            }
        )

    df = pd.DataFrame(grouped_rows)
    source = df["My Deck Winrate Source"].copy()
    df = df[
        [
            "Deck",
            "Source Deck Names",
            "Meta",
            "Winrate",
            "Winrate Game Count",
            "Archetype",
            "My Deck Winrate",
            "My Deck Winrate Game Count",
        ]
    ].copy()
    df["Archetype"] = df["Archetype"].apply(normalize_archetype_label)
    df["My Deck Winrate Fallback180"] = source.eq("fallback")
    df["My Deck Winrate Imputed"] = source.eq("none")
    df.loc[df["My Deck Winrate Imputed"], "My Deck Winrate"] = 0.5
    df = df.sort_values("Meta", ascending=False).reset_index(drop=True)
    for col in ["Meta", "Winrate", "My Deck Winrate"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").round(2)
    stats = {
        "primary": int(source.eq("primary").sum()),
        "fallback": int(source.eq("fallback").sum()),
        "imputed": int(source.eq("none").sum()),
    }
    return df, mapped_from_user, stats


def _fallback_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def _safe_to_csv(df: pd.DataFrame, path: Path) -> Path:
    final_path = path
    try:
        df.to_csv(final_path, index=False, encoding="utf-8")
    except PermissionError:
        final_path = _fallback_path(path)
        df.to_csv(final_path, index=False, encoding="utf-8")
    return final_path


def _remove_if_exists(path: Path) -> None:
    try:
        if path.exists():
            path.unlink()
    except Exception:
        pass


def _clear_previous_run_outputs(run_dir: Path, clear_challenge_outputs: bool = True) -> None:
    if not run_dir.exists():
        return

    for child in run_dir.iterdir():
        try:
            if not clear_challenge_outputs:
                if child.is_dir() and child.name == "statistics":
                    challenge_dir = child / "challenges"
                    if challenge_dir.exists() and challenge_dir.is_dir():
                        # Keep challenge statistics; clear only non-challenge statistics.
                        for stats_child in child.iterdir():
                            if stats_child.resolve() == challenge_dir.resolve():
                                continue
                            if stats_child.is_dir():
                                shutil.rmtree(stats_child, ignore_errors=True)
                            else:
                                stats_child.unlink(missing_ok=True)
                        continue
                if child.name.startswith("challenge_"):
                    continue

            if child.is_dir():
                shutil.rmtree(child, ignore_errors=True)
            else:
                child.unlink()
        except Exception:
            pass


_RUN_DIR_RE = re.compile(r"^W\d{2,}_(\d{4}-\d{2}-\d{2})_to_(\d{4}-\d{2}-\d{2})$")


def compute_run_dir(outputs_base: Path, week_start: date, week_end: date) -> Path:
    """Returns the canonical run directory for a given week window.

    Layout: outputs_base / YYYY / MM / WNN_YYYY-MM-DD_to_YYYY-MM-DD
    The WNN prefix uses the ISO week number of week_end for natural sorting.
    """
    iso_week = week_end.isocalendar()[1]
    folder = f"W{iso_week:02d}_{week_start.isoformat()}_to_{week_end.isoformat()}"
    return outputs_base / f"{week_end.year:04d}" / f"{week_end.month:02d}" / folder


def scan_run_dirs(outputs_base: Path) -> List[tuple]:
    """Scans outputs_base for all run directories matching the WNN_date_to_date pattern.

    Returns a sorted list of (week_start, week_end, run_dir) tuples, oldest first.
    """
    results: List[tuple] = []
    base = Path(outputs_base)
    if not base.exists():
        return results
    for year_dir in sorted(base.iterdir()):
        if not year_dir.is_dir() or not year_dir.name.isdigit():
            continue
        for month_dir in sorted(year_dir.iterdir()):
            if not month_dir.is_dir() or not month_dir.name.isdigit():
                continue
            for run_dir in sorted(month_dir.iterdir()):
                if not run_dir.is_dir():
                    continue
                m = _RUN_DIR_RE.match(run_dir.name)
                if not m:
                    continue
                try:
                    ws = date.fromisoformat(m.group(1))
                    we = date.fromisoformat(m.group(2))
                    results.append((ws, we, run_dir))
                except ValueError:
                    continue
    results.sort(key=lambda x: x[0])
    return results


def export_outputs(
    df: pd.DataFrame,
    output_csv: Path,
    output_xlsx: Path,
    include_csv: bool = True,
    include_xlsx: bool = True,
) -> tuple[Optional[Path], Optional[Path]]:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    export_df = df.drop(columns=["My Deck Winrate Imputed", "My Deck Winrate Fallback180"], errors="ignore").copy()
    for col in ["Meta", "Winrate", "My Deck Winrate"]:
        if col in export_df.columns:
            export_df[col] = pd.to_numeric(export_df[col], errors="coerce").round(2)
    imputed_flags = df.get("My Deck Winrate Imputed", pd.Series([False] * len(df))).astype(bool).tolist()
    fallback180_flags = df.get("My Deck Winrate Fallback180", pd.Series([False] * len(df))).astype(bool).tolist()

    final_csv: Optional[Path] = None
    final_xlsx: Optional[Path] = None

    if include_csv:
        try:
            final_csv = _safe_to_csv(export_df, output_csv)
        except Exception:
            raise

    yellow_fill = PatternFill(fill_type="solid", start_color="FFF59D", end_color="FFF59D")
    green_fill = PatternFill(fill_type="solid", start_color="C8E6C9", end_color="C8E6C9")
    wr_col_idx = export_df.columns.get_loc("My Deck Winrate") + 1

    if include_xlsx:
        final_xlsx = output_xlsx
        try:
            with pd.ExcelWriter(final_xlsx, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False)
                ws = writer.sheets["Sheet1"]
                for row_idx, is_imputed in enumerate(imputed_flags, start=2):
                    if is_imputed:
                        ws.cell(row=row_idx, column=wr_col_idx).fill = yellow_fill
                for row_idx, is_fallback in enumerate(fallback180_flags, start=2):
                    if is_fallback:
                        ws.cell(row=row_idx, column=wr_col_idx).fill = green_fill
        except PermissionError:
            final_xlsx = _fallback_path(output_xlsx)
            with pd.ExcelWriter(final_xlsx, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False)
                ws = writer.sheets["Sheet1"]
                for row_idx, is_imputed in enumerate(imputed_flags, start=2):
                    if is_imputed:
                        ws.cell(row=row_idx, column=wr_col_idx).fill = yellow_fill
                for row_idx, is_fallback in enumerate(fallback180_flags, start=2):
                    if is_fallback:
                        ws.cell(row=row_idx, column=wr_col_idx).fill = green_fill

    return final_csv, final_xlsx


def aggregate_by_deck(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    required = {"Deck", "Meta", "Winrate", "Archetype"}
    if not required.issubset(set(out.columns)):
        return out

    def merge_archetype(series: pd.Series) -> str:
        known = [
            normalize_archetype_label(s)
            for s in series.astype(str).tolist()
            if s and normalize_archetype_label(s) != "Unknown"
        ]
        return known[0] if known else "Unknown"

    def merge_my_wr(series: pd.Series, meta_series: pd.Series):
        mask = pd.Series(series).notna()
        if not mask.any():
            return pd.NA
        wr_valid = pd.Series(series)[mask].astype(float)
        meta_valid = pd.Series(meta_series)[mask].astype(float)
        total_meta = meta_valid.sum()
        if total_meta > 0:
            return (wr_valid * meta_valid).sum() / total_meta
        return wr_valid.mean()

    def merge_count(series: pd.Series) -> int:
        return int(pd.to_numeric(series, errors="coerce").fillna(0).sum())

    grouped = (
        out.groupby("Deck", as_index=False)
        .apply(
            lambda g: pd.Series(
                {
                    "Source Deck Names": " | ".join(sorted(set(g["Source Deck Names"].astype(str).tolist()))),
                    "Meta": pd.to_numeric(g["Meta"], errors="coerce").fillna(0).sum(),
                    "Winrate": (
                        (pd.to_numeric(g["Winrate"], errors="coerce").fillna(0) * pd.to_numeric(g["Meta"], errors="coerce").fillna(0)).sum()
                        / pd.to_numeric(g["Meta"], errors="coerce").fillna(0).sum()
                        if pd.to_numeric(g["Meta"], errors="coerce").fillna(0).sum() > 0
                        else pd.to_numeric(g["Winrate"], errors="coerce").dropna().mean()
                    ),
                    "Winrate Game Count": merge_count(g.get("Winrate Game Count", pd.Series(dtype=float))),
                    "Archetype": merge_archetype(g["Archetype"]),
                    "My Deck Winrate": merge_my_wr(g.get("My Deck Winrate", pd.Series(dtype=float)), g["Meta"]),
                    "My Deck Winrate Game Count": merge_count(g.get("My Deck Winrate Game Count", pd.Series(dtype=float))),
                    "My Deck Winrate Fallback180": bool(g.get("My Deck Winrate Fallback180", pd.Series(dtype=bool)).fillna(False).any()),
                    "My Deck Winrate Imputed": bool(g.get("My Deck Winrate Imputed", pd.Series(dtype=bool)).fillna(False).any()),
                }
            )
        )
        .reset_index(drop=True)
    )

    grouped = grouped.sort_values("Meta", ascending=False).reset_index(drop=True)
    return grouped


def apply_rogue_threshold_to_grouped(df: pd.DataFrame, rogue_threshold: float) -> pd.DataFrame:
    out = df.copy()
    if out.empty or "Meta" not in out.columns:
        return out

    out["Meta Tier"] = "Main"
    meta = pd.to_numeric(out["Meta"], errors="coerce")
    low_meta_mask = meta.notna() & (meta < float(rogue_threshold))
    out.loc[low_meta_mask, "Meta Tier"] = "Rogue"
    return out


def aggregate_rogue_bucket(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "Meta Tier" in out.columns:
        tier_series = out["Meta Tier"].fillna("").astype(str).str.strip().str.lower()
        rogue_mask = tier_series.eq("rogue")
    else:
        archetype_series = out.get("Archetype", pd.Series(dtype=str)).fillna("").astype(str).str.strip().str.lower()
        rogue_mask = archetype_series.isin({"rogue", "other"})

    if not rogue_mask.any():
        if "Meta Tier" not in out.columns:
            out["Meta Tier"] = "Main"
        return out

    rogue_rows = out.loc[rogue_mask].copy()
    keep_rows = out.loc[~rogue_mask].copy()
    if "Meta Tier" not in keep_rows.columns:
        keep_rows["Meta Tier"] = "Main"

    meta_sum = float(pd.to_numeric(rogue_rows["Meta"], errors="coerce").fillna(0).sum())
    winrate_vals = pd.to_numeric(rogue_rows["Winrate"], errors="coerce")
    meta_vals = pd.to_numeric(rogue_rows["Meta"], errors="coerce").fillna(0)
    if meta_sum > 0:
        winrate_agg = float((winrate_vals.fillna(0) * meta_vals).sum() / meta_sum)
    else:
        winrate_agg = float(winrate_vals.dropna().mean()) if winrate_vals.notna().any() else 0.5

    my_wr_vals = pd.to_numeric(rogue_rows["My Deck Winrate"], errors="coerce")
    my_wr_mask = my_wr_vals.notna()
    if my_wr_mask.any():
        my_meta = meta_vals[my_wr_mask]
        if my_meta.sum() > 0:
            my_wr_agg = float((my_wr_vals[my_wr_mask] * my_meta).sum() / my_meta.sum())
        else:
            my_wr_agg = float(my_wr_vals[my_wr_mask].mean())
    else:
        my_wr_agg = 0.5

    source_names = sorted(
        set(rogue_rows["Source Deck Names"].dropna().astype(str).tolist())
        | set(rogue_rows["Deck"].dropna().astype(str).tolist())
    )

    rogue_row = {
        "Deck": "Rogue",
        "Source Deck Names": " | ".join(source_names),
        "Meta": meta_sum,
        "Winrate": winrate_agg,
        "Winrate Game Count": int(pd.to_numeric(rogue_rows["Winrate Game Count"], errors="coerce").fillna(0).sum()),
        "Archetype": "Mixed",
        "Meta Tier": "Rogue",
        "My Deck Winrate": my_wr_agg,
        "My Deck Winrate Game Count": int(pd.to_numeric(rogue_rows["My Deck Winrate Game Count"], errors="coerce").fillna(0).sum()),
        "My Deck Winrate Fallback180": bool(rogue_rows.get("My Deck Winrate Fallback180", pd.Series(dtype=bool)).fillna(False).any()),
        "My Deck Winrate Imputed": bool(rogue_rows.get("My Deck Winrate Imputed", pd.Series(dtype=bool)).fillna(False).any()),
    }

    combined = pd.concat([keep_rows, pd.DataFrame([rogue_row])], ignore_index=True)
    combined = combined.sort_values("Meta", ascending=False).reset_index(drop=True)
    return combined


def export_unknown_archetypes(df: pd.DataFrame, path: Path) -> Path:
    unknown = df[df["Archetype"] == "Unknown"][["Deck", "Source Deck Names"]].drop_duplicates().sort_values("Deck")
    path.parent.mkdir(parents=True, exist_ok=True)
    if len(unknown) == 0:
        return _safe_to_csv(pd.DataFrame(columns=["Deck", "Source Deck Names", "Suggested Archetype"]), path)
    unknown = unknown.assign(**{"Suggested Archetype": ""})
    return _safe_to_csv(unknown, path)


def export_alias_suggestions(df: pd.DataFrame, aliases: List[DeckAliasRule], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    unknown = df[df["Archetype"] == "Unknown"].copy()
    if len(unknown) == 0:
        return _safe_to_csv(
            pd.DataFrame(columns=["Deck", "Suggested Canonical Name", "Confidence", "Reason"]),
            path,
        )

    canonical_pool = sorted(set([a.canonical_name for a in aliases] + df[df["Archetype"] != "Unknown"]["Deck"].astype(str).tolist()))
    results = []

    for deck in unknown["Deck"].astype(str).tolist():
        norm = normalize_name(deck)
        if re.fullmatch(r"[wubrg]{1,5}", norm):
            results.append(
                {
                    "Deck": deck,
                    "Suggested Canonical Name": "",
                    "Confidence": 0.0,
                    "Reason": "Color code placeholder (manual decision)",
                }
            )
            continue

        match = difflib.get_close_matches(deck, canonical_pool, n=1, cutoff=0.74)
        if not match:
            results.append(
                {
                    "Deck": deck,
                    "Suggested Canonical Name": "",
                    "Confidence": 0.0,
                    "Reason": "No close canonical match",
                }
            )
            continue

        candidate = match[0]
        confidence = difflib.SequenceMatcher(a=normalize_name(deck), b=normalize_name(candidate)).ratio()
        results.append(
            {
                "Deck": deck,
                "Suggested Canonical Name": candidate,
                "Confidence": round(confidence, 3),
                "Reason": "Name similarity",
            }
        )

    result_df = pd.DataFrame(results).sort_values(["Confidence", "Deck"], ascending=[False, True])
    return _safe_to_csv(result_df, path)


def _fmt_num(value, decimals: int = 2) -> str:
    if value is None or pd.isna(value):
        return ""
    try:
        return f"{float(value):.{decimals}f}"
    except Exception:
        return str(value)


def export_grouped_xml(df: pd.DataFrame, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    root = ET.Element("metagame")
    root.set("generated_at", datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"))
    root.set("rows", str(len(df)))

    for archetype, group in df.groupby("Archetype", dropna=False, sort=True):
        archetype_name = str(archetype or "Unknown")
        archetype_node = ET.SubElement(root, "archetype")
        archetype_node.set("name", archetype_name)
        archetype_node.set("meta", _fmt_num(pd.to_numeric(group["Meta"], errors="coerce").fillna(0).sum(), 4))
        archetype_node.set("decks", str(len(group)))

        ordered = group.sort_values("Meta", ascending=False)
        for _, row in ordered.iterrows():
            deck_node = ET.SubElement(archetype_node, "deck")
            deck_node.set("name", str(row.get("Deck") or ""))
            deck_node.set("source_names", str(row.get("Source Deck Names") or ""))
            deck_node.set("meta", _fmt_num(row.get("Meta"), 4))
            deck_node.set("winrate", _fmt_num(row.get("Winrate"), 4))
            deck_node.set("winrate_game_count", str(int(pd.to_numeric(row.get("Winrate Game Count"), errors="coerce") or 0)))
            deck_node.set("my_deck_winrate", _fmt_num(row.get("My Deck Winrate"), 4))
            deck_node.set(
                "my_deck_winrate_game_count",
                str(int(pd.to_numeric(row.get("My Deck Winrate Game Count"), errors="coerce") or 0)),
            )

    tree = ET.ElementTree(root)
    try:
        ET.indent(tree, space="  ")
    except Exception:
        pass
    tree.write(path, encoding="utf-8", xml_declaration=True)
    return path


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Buduje plik wejściowy do MTG-Metagame-Analyzer: tygodniowy metagame + "
            "My Deck Winrate z dłuższego okna matchupów (np. Domain Zoo)."
        )
    )
    parser.add_argument("--format", dest="format_name", default="Modern")
    parser.add_argument("--week-start", help="YYYY-MM-DD (tryb ręczny)")
    parser.add_argument("--week-end", help="YYYY-MM-DD (tryb ręczny)")
    parser.add_argument("--metagame-window-days", type=int, default=14, help="okno metagame, domyślnie 14 dni")
    parser.add_argument("--history-points", type=int, default=1, help="ile kolejnych punktów tygodniowych wygenerować")
    parser.add_argument("--anchor-sunday", help="YYYY-MM-DD, domyślnie ostatnia niedziela")
    parser.add_argument("--history-output-dir", default="outputs/history", help="(deprecated – use --outputs-base)")
    parser.add_argument("--outputs-base", default=None, help="katalog bazowy; jeśli pominięty – wyprowadzany z --history-output-dir")
    parser.add_argument("--my-deck", default="Domain Zoo")
    parser.add_argument("--my-window-days", type=int, default=90, help="np. 30 lub 90")
    parser.add_argument("--my-fallback-window-days", type=int, default=180, help="okno fallback dla brakow, np. 180")
    parser.add_argument("--rogue-threshold", type=float, default=0.5, help="próg Meta (%%) poniżej którego deck wpada do Rogue")
    parser.add_argument("--metagame-limit", type=int, default=64)
    parser.add_argument("--matchup-limit", type=int, default=300)
    parser.add_argument(
        "--rules-file",
        default="docs/archetype_rules.csv",
        help="CSV z regułami mapowania archetypów",
    )
    parser.add_argument(
        "--aliases-file",
        default="docs/deck_aliases.csv",
        help="CSV z aliasami nazw decków do nazw kanonicznych",
    )
    parser.add_argument(
        "--user-mapping-file",
        default="docs/user_deck_mapping.csv",
        help="Twoj plik mapowan: raw_name, canonical_name, archetype",
    )
    parser.add_argument(
        "--challenge-mapping-file",
        default="docs/challenge_deck_mapping.csv",
        help="Osobny plik mapowan dla danych Challenge: raw_name, canonical_name, archetype",
    )
    parser.add_argument(
        "--output-csv",
        default="outputs/metagame_input.csv",
    )
    parser.add_argument(
        "--output-xlsx",
        default="outputs/metagame_input.xlsx",
    )
    parser.add_argument(
        "--output-csv-rogue-grouped",
        default="outputs/metagame_input_rogue_grouped.csv",
    )
    parser.add_argument(
        "--output-xlsx-rogue-grouped",
        default="outputs/metagame_input_rogue_grouped.xlsx",
    )
    parser.add_argument(
        "--output-xml-grouped",
        default="outputs/metagame_input_grouped.xml",
    )
    parser.add_argument(
        "--unknown-output",
        default="outputs/unknown_archetypes.csv",
    )
    parser.add_argument(
        "--alias-suggestions-output",
        default="outputs/alias_suggestions.csv",
    )
    parser.add_argument(
        "--output-profile",
        choices=["full", "analysis"],
        default="full",
        help="full: all outputs, analysis: base CSV + grouped XLSX",
    )
    parser.add_argument(
        "--include-challenge-decklist",
        action="store_true",
        help="Pobiera wszystkie Challenge (C32+C64) z MTGGoldfish w oknie metagame i zapisuje raporty + historię + statystyki.",
    )
    parser.add_argument(
        "--challenge-history-events",
        type=int,
        default=8,
        metavar="N",
        help="Ile ostatnich eventów (per rozmiar) użyć do obliczania statystyk historii Challenge (default: 8).",
    )
    parser.add_argument(
        "--backup-url",
        default="",
        help="Stały URL backup API; może wskazywać host root albo /api.",
    )
    parser.add_argument(
        "--allow-partial",
        action="store_true",
        help=(
            "Gdy okno Challenge jest niekompletne (rejestr mtgo.com != dane pobrane), zapisz mimo to "
            "challenge_statistics_PARTIAL.xlsx zamiast pomijać zapis (domyślnie: nie zapisuj)."
        ),
    )
    return parser.parse_args(argv)


def resolve_windows(args: argparse.Namespace) -> List[tuple[date, date]]:
    if args.history_points < 1:
        raise ValueError("history-points musi być >= 1")

    if args.week_start or args.week_end:
        if not (args.week_start and args.week_end):
            raise ValueError("Podaj oba parametry: --week-start i --week-end")
        start = parse_date(args.week_start)
        end = parse_date(args.week_end)
        if end < start:
            raise ValueError("week-end musi być >= week-start")
        return [(start, end)]

    anchor = parse_date(args.anchor_sunday) if args.anchor_sunday else latest_sunday(date.today())
    windows: List[tuple[date, date]] = []
    for i in range(args.history_points):
        weeks_back = args.history_points - 1 - i
        end = anchor - timedelta(days=7 * weeks_back)
        start = end - timedelta(days=max(1, args.metagame_window_days) - 1)
        windows.append((start, end))
    return windows


def interactive_menu() -> argparse.Namespace:
    """Interaktywne menu zamiast linii komend"""
    print("\n" + "="*50)
    print("  MTG Metagame Generator - Menu")
    print("="*50 + "\n")
    
    # Pytanie o format
    print("Dostępne formaty:")
    formats = ["Modern", "Legacy", "Pioneer", "Pauper", "Standard", "Commander"]
    for i, fmt in enumerate(formats, 1):
        print(f"  {i}. {fmt}")
    
    while True:
        try:
            choice = input("\nWybierz format (1-6) [1]: ").strip()
            if not choice:
                format_choice = "Modern"
                break
            choice_num = int(choice)
            if 1 <= choice_num <= len(formats):
                format_choice = formats[choice_num - 1]
                break
            print("Błędny wybór. Spróbuj ponownie.")
        except ValueError:
            print("Wpisz numer (1-6).")
    
    # Pytanie o deck
    print(f"\nWybrano: {format_choice}")
    my_deck = input("Podaj nazwę Twojego decka [Domain Zoo]: ").strip()
    if not my_deck:
        my_deck = "Domain Zoo"
    
    print(f"\nTwój deck: {my_deck}")
    print("\nUruchamiam generator...\n")
    
    # Zwróć args w tym samym formacie co parse_args()
    args = parse_args([])
    args.format_name = format_choice
    args.my_deck = my_deck
    return args


def _write_challenge_failure_status(run_dir: Path, week_start: date, week_end: date, reason: str, log) -> None:
    """Fail-closed status.json for the case where the Challenge step blew up before it could even
    build a completeness signal (registry fetch threw, sync threw, anything unexpected). Every
    count is null (never 0, never silently "complete") -- same shape as
    challenge_mtgo_source.unavailable_completeness_summary(), duplicated here (not imported) to
    avoid this module depending on challenge_mtgo_source at import time.
    """
    status_dir = run_dir / "statistics" / "challenges"
    try:
        status_dir.mkdir(parents=True, exist_ok=True)
        payload = {
            "complete": False,
            "registry_count": None,
            "fetched_count": None,
            "missing": [],
            "window_start": week_start.isoformat() if week_start else None,
            "window_end": week_end.isoformat() if week_end else None,
            "league_window_start": None,
            "league_window_end": None,
            "window_consistent": None,
            "tier_counts_registry": {},
            "tier_counts_fetched": {},
            "premier_checked": False,
            "premier_count": None,
            "premier_events": [],
            "premier_note": "Premier scan did not complete",
            "partial": False,
            "failure_reason": reason,
        }
        (status_dir / "challenge_statistics.status.json").write_text(
            json.dumps(payload, indent=2), encoding="utf-8"
        )
        log(f"[ERROR] Wrote fail-closed status.json (complete=false) -> {status_dir / 'challenge_statistics.status.json'}")
    except Exception as write_err:
        log(f"[ERROR] Could not even write the fail-closed status.json: {write_err}")


def run_generation(args: argparse.Namespace, log=print) -> List[GenerationRunResult]:
    configure_backup_server_url(getattr(args, "backup_url", ""))
    windows = resolve_windows(args)

    rules = load_rules(Path(args.rules_file))
    aliases = load_aliases(Path(args.aliases_file))
    user_mappings = load_user_deck_mappings(Path(args.user_mapping_file))
    challenge_extra_mappings = load_user_deck_mappings(Path(getattr(args, "challenge_mapping_file", "docs/challenge_deck_mapping.csv")))
    # Keep naming consistent across Meta and Challenge: base mapping is shared.
    # Challenge mapping file is treated as optional additive extension only.
    challenge_mappings = list(user_mappings)
    known_raw = {normalize_name(m.raw_name) for m in challenge_mappings if normalize_name(m.raw_name)}
    for mapping in challenge_extra_mappings:
        key = normalize_name(mapping.raw_name)
        if key and key not in known_raw:
            challenge_mappings.append(mapping)
            known_raw.add(key)
    user_mapping = mapping_lookup(user_mappings)
    output_profile = getattr(args, "output_profile", "full")
    analysis_mode = output_profile == "analysis"

    # Resolve outputs base directory from --outputs-base or fall back to parent of --history-output-dir
    if getattr(args, "outputs_base", None):
        outputs_base = Path(args.outputs_base)
    else:
        outputs_base = Path(args.history_output_dir).parent

    log(f"[OK] User mapping rows loaded: {len(user_mappings)}")
    if challenge_extra_mappings:
        log(
            f"[OK] Challenge extra mapping rows loaded: {len(challenge_extra_mappings)} "
            f"(effective shared+extra: {len(challenge_mappings)})"
        )
    results: List[GenerationRunResult] = []

    for idx, (week_start, week_end) in enumerate(windows, start=1):
        dataset, mapped_from_user, my_wr_stats = build_dataset(
            format_name=args.format_name,
            week_start=week_start,
            week_end=week_end,
            my_deck=args.my_deck,
            my_window_days=args.my_window_days,
            my_fallback_window_days=args.my_fallback_window_days,
            rogue_threshold=args.rogue_threshold,
            rules=rules,
            aliases=aliases,
            user_mapping=user_mapping,
            metagame_limit=args.metagame_limit,
            matchup_limit=args.matchup_limit,
        )

        run_dir = compute_run_dir(outputs_base, week_start, week_end)
        run_dir.mkdir(parents=True, exist_ok=True)
        _clear_previous_run_outputs(
            run_dir,
            clear_challenge_outputs=bool(getattr(args, "include_challenge_decklist", False)),
        )

        output_csv = run_dir / "metagame_input.csv"
        output_xlsx = run_dir / "metagame_input.xlsx"
        output_csv_grouped = run_dir / "metagame_input_rogue_grouped.csv"
        output_xlsx_grouped = run_dir / "metagame_input_rogue_grouped.xlsx"
        output_xml_grouped = run_dir / "metagame_input_grouped.xml"
        unknown_output = run_dir / "unknown_archetypes.csv"
        alias_suggestions_output = run_dir / "alias_suggestions.csv"

        if analysis_mode:
            output_xlsx_grouped = run_dir / "metagame_input_grouped.xlsx"
            _remove_if_exists(output_xlsx)
            _remove_if_exists(output_csv_grouped)
            _remove_if_exists(output_xml_grouped)
            _remove_if_exists(run_dir / "metagame_input_rogue_grouped.xlsx")
            _remove_if_exists(unknown_output)
            _remove_if_exists(alias_suggestions_output)

        final_csv, final_xlsx = export_outputs(
            dataset,
            output_csv=output_csv,
            output_xlsx=output_xlsx,
            include_csv=True,
            include_xlsx=not analysis_mode,
        )
        grouped = aggregate_by_deck(dataset)
        grouped = apply_rogue_threshold_to_grouped(grouped, args.rogue_threshold)
        grouped_csv, grouped_xlsx = export_outputs(
            grouped,
            output_csv=output_csv_grouped,
            output_xlsx=output_xlsx_grouped,
            include_csv=not analysis_mode,
            include_xlsx=True,
        )
        grouped_xml = None
        final_unknown = None
        final_alias_suggestions = None
        challenge_decklist_path: Optional[Path] = None
        challenge_compare_path: Optional[Path] = None
        challenge_event_url: Optional[str] = None
        challenge_history_path: Optional[Path] = None
        premier_history_path: Optional[Path] = None
        challenge_stats_path: Optional[Path] = None
        league_results_path: Optional[Path] = None
        metagame_stats_path: Optional[Path] = None
        challenge_events_fetched: int = 0
        challenge_review_items: List[dict] = []

        if not analysis_mode:
            grouped_xml = export_grouped_xml(grouped, output_xml_grouped)
            final_unknown = export_unknown_archetypes(dataset, path=unknown_output)
            final_alias_suggestions = export_alias_suggestions(dataset, aliases=aliases, path=alias_suggestions_output)

        if bool(getattr(args, "include_challenge_decklist", False)):
            try:
                from challenge_mtgo_source import (
                    ChallengeSourceError,
                    build_challenge_dataset,
                    write_challenge_manifest,
                    write_event_status_manifest,
                    write_premier_events_csv,
                    write_review_items_csv,
                )

                _format_slug_hist = _format_slug_for_url(args.format_name)
                cache_dir = outputs_base / "cache" / "mtgo_json" / _format_slug_hist
                challenge_history_csv = outputs_base / f"challenge_history_{_format_slug_hist}.csv"
                premier_history_csv = outputs_base / f"premier_history_{_format_slug_hist}.csv"

                challenge_dataset = build_challenge_dataset(
                    format_name=args.format_name,
                    start_date=week_start,
                    end_date=week_end,
                    aliases=aliases,
                    rules=rules,
                    challenge_mappings=challenge_mappings,
                    cache_dir=cache_dir,
                    log=log,
                    challenge_history_csv=challenge_history_csv,
                    premier_history_csv=premier_history_csv,
                )
                challenge_events_fetched = len(challenge_dataset.challenge_events)
                log(
                    f"[OK] Fetched {challenge_events_fetched} challenge event(s) and "
                    f"{len(challenge_dataset.premier_events)} premier event(s) from mtgo.com"
                )

                window_label = f"{week_start.isoformat()}_to_{week_end.isoformat()}"
                write_challenge_manifest(challenge_dataset, run_dir / f"challenge_manifest_{window_label}.csv", log=log)
                write_premier_events_csv(
                    challenge_dataset, _format_slug_hist, run_dir / f"premier_events_{window_label}.csv", log=log
                )

                # Completeness gate: registry_count (everything mtgo.com listed for the window) vs
                # fetched_count (what actually got parsed). Never narrows the window to match what
                # was fetched -- a shortfall is reported (event manifest + log + stats stamp), not hidden.
                completeness = write_event_status_manifest(
                    challenge_dataset, run_dir / f"challenge_event_status_{window_label}.csv", log=log,
                    format_name=args.format_name,
                )
                if not completeness.complete:
                    log(
                        f"[WARN] Challenge window INCOMPLETE: {completeness.fetched_count}/"
                        f"{completeness.registry_count} event(s) fetched for "
                        f"{week_start.isoformat()}..{week_end.isoformat()}; missing "
                        f"{[(m.event_id, m.date, m.status) for m in completeness.missing]}"
                    )
                if challenge_dataset.review_items:
                    review_path = write_review_items_csv(
                        challenge_dataset, run_dir / f"challenge_needs_review_{window_label}.csv", log=log
                    )
                    log(
                        f"[WARN] {len(challenge_dataset.review_items)} deck(s) could not be confidently "
                        f"labeled -> {review_path}"
                    )
                    challenge_review_items = [
                        {
                            "event_id": item.event_id,
                            "date": item.date,
                            "format": item.format_name,
                            "kind": item.kind,
                            "tier": item.size,
                            "place": item.place,
                            "pilot": item.pilot,
                            "reason": item.reason,
                            "nearest_label": item.nearest_label or "",
                            "nearest_sim": item.nearest_sim,
                            "second_label": item.second_label or "",
                            "second_sim": item.second_sim,
                            "key_cards": ", ".join(f"{name} x{qty}" for name, qty in item.key_cards),
                            "event_url": item.url,
                        }
                        for item in challenge_dataset.review_items
                    ]

                for event in challenge_dataset.challenge_events:
                    ev_rows = challenge_dataset.event_rows.get(event.event_id, [])
                    ch_df = pd.DataFrame(
                        [{"Place": r.place, "Deck": r.deck, "Archetype": r.archetype, "Pilot": r.pilot} for r in ev_rows]
                    ).sort_values("Place")
                    c_label = f"C{event.size}_{event.date}_{event.event_id}"

                    ch_csv_path = _safe_to_csv(ch_df, run_dir / f"challenge_{c_label}_decklist.csv")
                    if challenge_decklist_path is None:
                        challenge_decklist_path = ch_csv_path
                    if challenge_event_url is None:
                        challenge_event_url = event.url

                    compare_df = build_challenge_vs_metagame_report(
                        challenge_df=ch_df,
                        metagame_df=dataset,
                        aliases=aliases,
                        rules=rules,
                    )
                    cmp_csv = _safe_to_csv(compare_df, run_dir / f"challenge_{c_label}_vs_metagame.csv")
                    if challenge_compare_path is None:
                        challenge_compare_path = cmp_csv

                    log(f"[OK] Challenge {c_label}: {len(ch_df)} decks -> {ch_csv_path}")

                # History CSV stored at repo level (not inside a weekly run dir); the window is
                # synced (old rows for these dates purged, fresh EventID-keyed rows inserted), not
                # appended -- so re-running a window can never leave stale/contaminated rows behind.
                if sync_challenge_history_window is not None:
                    sync_challenge_history_window(
                        history_csv=challenge_history_csv,
                        format_name=args.format_name,
                        start_date=week_start,
                        end_date=week_end,
                        dataset=challenge_dataset,
                        log=log,
                    )
                    challenge_history_path = challenge_history_csv

                # Premier events never enter the Challenge history/stats, but they get their own
                # parallel history file (same EventID-keyed sync) so manual review corrections for
                # them have somewhere permanent to land -- not just the per-run premier_events csv.
                if sync_challenge_history_window is not None and challenge_dataset.premier_events:
                    premier_dataset_view = SimpleNamespace(
                        challenge_events=challenge_dataset.premier_events,
                        event_rows=challenge_dataset.premier_rows,
                    )
                    sync_challenge_history_window(
                        history_csv=premier_history_csv,
                        format_name=args.format_name,
                        start_date=week_start,
                        end_date=week_end,
                        dataset=premier_dataset_view,
                        log=log,
                    )
                    premier_history_path = premier_history_csv

                if (
                    run_challenge_statistics is not None
                    and challenge_history_path is not None
                    and not getattr(args, "skip_challenge_stats", False)
                ):
                    last_n = int(getattr(args, "challenge_history_events", 8))
                    challenge_stats_dir = run_dir / "statistics" / "challenges"
                    premier_event_ids = {e.event_id for e in challenge_dataset.premier_events}
                    ch_stats_result = run_challenge_statistics(
                        history_csv=challenge_history_path,
                        output_dir=challenge_stats_dir,
                        format_name=args.format_name,
                        last_n_events=last_n,
                        week_start=week_start,
                        week_end=week_end,
                        metagame_df=dataset,
                        log=log,
                        premier_event_ids=premier_event_ids,
                        completeness=completeness,
                        allow_partial=bool(getattr(args, "allow_partial", False)),
                    )
                    challenge_stats_path = ch_stats_result.excel_path

                if run_league_update is not None and challenge_history_path is not None:
                    league_dir = outputs_base / "league"
                    league_summary = run_league_update(
                        history_csv=challenge_history_path,
                        league_dir=league_dir,
                        format_name=args.format_name,
                        start_date=week_start,
                        end_date=week_end,
                        as_of=date.today(),
                        premier_history_csv=premier_history_csv,
                        mtgo_json_cache_dir=cache_dir,
                        log=log,
                    )
                    if league_summary["written_event_ids"]:
                        league_results_path = league_dir / "results"

                    if export_league_site is not None:
                        try:
                            export_league_site(
                                league_dir=league_dir,
                                docs_data_dir=Path("docs") / "data",
                                as_of=date.today(),
                                log=log,
                            )
                        except Exception as site_err:
                            # Site export is a read-only downstream view of data already durably
                            # written above -- a failure here must not be mistaken for a league
                            # sync failure (which does abort further processing, see the except
                            # blocks below) and must not stop the rest of the weekly run either.
                            log(f"[ERROR] [league-site] site data export FAILED (league data itself is unaffected): {site_err}")

            except ChallengeSourceError as challenge_err:
                log(f"[ERROR] Challenge report STOPPED (not a silent skip): {challenge_err}")
                _write_challenge_failure_status(run_dir, week_start, week_end, str(challenge_err), log)
            except Exception as challenge_err:
                # Deliberately NOT a silent skip: this is an *unexpected* failure (anything not
                # already raised as the structured ChallengeSourceError above), which previously
                # logged as a mere [WARN] and left no durable trace at all -- the same failure
                # pattern as the old "complete: true, registry_count: null" bug, just one level up
                # (no status.json ever written instead of one wrongly claiming success). It is
                # still caught here (a broken Challenge step must not abort the rest of the weekly
                # run -- metagame stats, etc. still need to run), but it is now always logged at
                # ERROR with a full traceback and always leaves a fail-closed status.json behind.
                log(f"[ERROR] Challenge report FAILED (not swallowed): {challenge_err}")
                log(traceback.format_exc())
                _write_challenge_failure_status(run_dir, week_start, week_end, str(challenge_err), log)

        # Generate metagame statistics
        if run_statistics is not None and grouped_xlsx is not None:
            try:
                metagame_stats_dir = run_dir / "statistics" / "metagame"
                metagame_stats_result = run_statistics(
                    input_excel=grouped_xlsx,
                    output_dir=metagame_stats_dir,
                    history_csv=outputs_base / f"metagame_history_{_format_slug_for_url(args.format_name)}.csv",
                    total_players=1000,
                    rounds=5,
                    min_encounter_pct=5.0,
                    player_deck_name=args.my_deck or "My Deck",
                    player_winrate=0.5,
                    weeks_back=4,
                    output_profile="full",
                    palette_name="classic",
                    deck_colors=None,
                    legend_colors=None,
                    log=log,
                )
                xlsx_files = [f for f in metagame_stats_result.files if f.suffix == ".xlsx"]
                metagame_stats_path = xlsx_files[0] if xlsx_files else metagame_stats_dir
            except Exception as metagame_err:
                log(f"[WARN] Metagame statistics skipped: {metagame_err}")

        log(
            f"[OK] P{idx} {week_start.isoformat()}..{week_end.isoformat()} | "
            f"rows={len(dataset)} | primary={my_wr_stats['primary']} | "
            f"180d={my_wr_stats['fallback']} | 50%={my_wr_stats['imputed']}"
        )
        log(f"[OK] Rogue threshold: Meta < {args.rogue_threshold}%")
        log(f"[OK] CSV: {final_csv}")
        if final_xlsx is not None:
            log(f"[OK] XLSX: {final_xlsx}")
        if grouped_csv is not None:
            log(f"[OK] CSV (Rogue grouped): {grouped_csv}")
        label = "grouped" if analysis_mode else "Rogue grouped"
        log(f"[OK] XLSX ({label}): {grouped_xlsx}")
        if grouped_xml is not None:
            log(f"[OK] XML (grouped): {grouped_xml}")
        if final_unknown is not None:
            log(f"[OK] Unknown archetypes report: {final_unknown}")
        if final_alias_suggestions is not None:
            log(f"[OK] Alias suggestions report: {final_alias_suggestions}")
        if challenge_decklist_path is not None:
            log(f"[OK] Challenge decklist(s) (MTGGoldfish): {challenge_decklist_path.parent}")
        if challenge_compare_path is not None:
            log(f"[OK] Challenge vs metagame: {challenge_compare_path.parent}")
        if challenge_event_url:
            log(f"[OK] Challenge source URL: {challenge_event_url}")
        if challenge_history_path is not None:
            log(f"[OK] Challenge history: {challenge_history_path}")
        if challenge_stats_path is not None:
            log(f"[OK] Challenge statistics: {challenge_stats_path}")
        if league_results_path is not None:
            log(f"[OK] League results: {league_results_path}")
        if metagame_stats_path is not None:
            log(f"[OK] Metagame statistics: {metagame_stats_path}")
        log(f"[OK] Output dir: {run_dir}")

        results.append(
            GenerationRunResult(
                week_start=week_start,
                week_end=week_end,
                run_dir=run_dir,
                row_count=len(dataset),
                mapped_from_user=mapped_from_user,
                primary_count=my_wr_stats["primary"],
                fallback_count=my_wr_stats["fallback"],
                imputed_count=my_wr_stats["imputed"],
                rogue_threshold=args.rogue_threshold,
                csv_path=final_csv,
                xlsx_path=final_xlsx,
                rogue_csv_path=grouped_csv,
                rogue_xlsx_path=grouped_xlsx,
                rogue_xml_path=grouped_xml,
                unknown_output_path=final_unknown,
                alias_suggestions_path=final_alias_suggestions,
                challenge_decklist_path=challenge_decklist_path,
                challenge_compare_path=challenge_compare_path,
                challenge_event_url=challenge_event_url,
                challenge_history_path=challenge_history_path,
                premier_history_path=premier_history_path,
                challenge_stats_path=challenge_stats_path,
                challenge_events_fetched=challenge_events_fetched,
                challenge_review_items=challenge_review_items,
            )
        )

    return results


def main() -> None:
    # Sprawdź czy uruchomiona z argumentami linii komend
    if len(sys.argv) > 1:
        args = parse_args()
    else:
        # Interaktywne menu
        args = interactive_menu()
    
    run_generation(args)


if __name__ == "__main__":
    try:
        main()
    except RuntimeError as err:
        print(f"\n[ERROR] {err}")
        sys.exit(2)
